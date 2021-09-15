#넥슨 닷컴 크롤러
from selenium import webdriver
import time
import openpyxl

options=webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches',['enable-logging'])
driver=webdriver.Chrome(options=options)
driver.get("https://www.nexon.com/Home/Game")

#엑셀 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title="정기정검"

url_list=[]
game_name_list=[]
games=driver.find_elements_by_css_selector("ul.pc.cardView")
games=games[0].find_elements_by_tag_name("li")
cnt=1
for i in games:
    game_a=i.find_element_by_class_name("gameName")
    a_tag=game_a.find_element_by_tag_name("a")
    url_list.append(a_tag.get_attribute("href"))
    game_name=driver.find_element_by_xpath("//*[@id='allGame']/ul[1]/li["+str(cnt)+"]/div/span[2]/a/span").text
    game_name_list.append(game_name)
    cnt+=1
sheet.cell(row=1,column=1).value="게임 이름"
sheet.cell(row=2,column=1).value="공식 사이트"
for i in range(len(url_list)):
    driver.get(url_list[i])
    sheet.cell(row=1,column=i+2).value=game_name_list[i]
    sheet.cell(row=2,column=i+2).value=url_list[i]
    time.sleep(0.3)






wb.save('Nexon_Crawler.xlsx')